import openpyxl

excel_file_path = "/Data Driven Testing/Test Excel Files/read_data.xlsx"
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook['Sheet1']
rows = sheet.max_row    # To get number of rows
columns = sheet.max_column  # To get number of cells
# Reading all the rows and columns from exel sheet
for row in range(1, rows + 1):
    for cell in range(1, columns + 1):
        print(sheet.cell(row, cell).value, end="      ")
    print()
