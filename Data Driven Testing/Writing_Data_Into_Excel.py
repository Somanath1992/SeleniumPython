import openpyxl

excel_file_path = "C:/Users/Admin/PycharmProjects/SeleniumPython/Data Driven Testing/Test Excel Files/write_data.xlsx"
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook.active  # If we have only one sheet and its active
# To enter same value in all rows and columns
for row in range(1, 6):
    for col in range(1, 4):
        sheet.cell(row, col).value = "welcome"

workbook.save(excel_file_path)  # Save the file after entering data

# To enter different data in rows and columns
excel_file_path1 = ("C:/Users/Admin/PycharmProjects/SeleniumPython/Data Driven Testing/Test Excel "
                    "Files/write_data_1.xlsx")
workbook = openpyxl.load_workbook(excel_file_path1)
sheet = workbook.active

sheet.cell(1, 1).value = 123
sheet.cell(1, 2).value = "John"
sheet.cell(2, 1).value = 456
sheet.cell(2, 2).value = "David"
sheet.cell(3, 1).value = 789
sheet.cell(3, 2).value = "Tom"
workbook.save(excel_file_path1)
