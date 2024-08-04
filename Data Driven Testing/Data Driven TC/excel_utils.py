import openpyxl
from openpyxl.styles import PatternFill


def get_row_count(excel_file_path, sheet_name):
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_col_count(excel_file_path, sheet_name):
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook[sheet_name]
    return sheet.max_column


def read_data(excel_file_path, sheet_name, row_num, col_num):
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook[sheet_name]
    return sheet.cell(row_num, col_num).value


def write_data(excel_file_path, sheet_name, row_num, col_num, data):
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook[sheet_name]
    sheet.cell(row_num, col_num).value = data
    workbook.save(excel_file_path)


def fill_green_color(excel_file_path, sheet_name, row_num, col_num):
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook[sheet_name]
    green_fill = PatternFill(start_color='60b212', end_color='60b212', fill_type='solid')
    sheet.cell(row_num, col_num).fill = green_fill
    workbook.save(excel_file_path)


def fill_red_color(excel_file_path, sheet_name, row_num, col_num):
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook[sheet_name]
    red_fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
    sheet.cell(row_num, col_num).fill = red_fill
    workbook.save(excel_file_path)

